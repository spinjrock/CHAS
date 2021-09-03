#Spencer Oswald 9/3/21
import openpyxl
#TODO Function descriptions
class SpreadsheetHandler:
    
    TOTALS = {}
    SPREADSHEET_NAME = ''
    WB = None
    
    def __init__(self):
        self.TOTALS = {'2021-09-02': 2}
        self.SPREADSHEET_NAME = 'default_spreadsheet.xlsx'
    
    def __init__(self, TOTALS: dict, SPREADSHEET_NAME: str):
        self.TOTALS = TOTALS
        self.SPREADSHEET_NAME = SPREADSHEET_NAME
        
    def get_spreadsheet(self):
        try:
            self.WB = openpyxl.load_workbook(self.SPREADSHEET_NAME)
        except:
            print("Creating workbooK: "+ self.SPREADSHEET_NAME)
            self.WB =openpyxl.Workbook()
    
    def format_spreadsheet(self):
        if self.WB != None:
            ws = self.WB.active
        elif self.WB == None:
            self.get_spreadsheet()
            ws = self.WB.active
        ws.delete_cols(1, 10)
        ws["A1"] = "Dates"
        ws["C1"] = "Daily Totals"
        #ws["E1"] = "Grand Total"
        for i in range(1, 364): #Doing some excessive cell merges
            ws.merge_cells("A"+str(i)+":B"+str(i))
            ws.merge_cells("C"+str(i)+":D"+str(i))
            ws.merge_cells("E"+str(i)+":F"+str(i))
        self.WB.save(self.SPREADSHEET_NAME)
    
    def format_money(self, money: int):
        return "$"+str(float(money) / 100)
    
    def write_spreadsheet(self):
        i = 2
        ws = self.WB.active
        for date in self.TOTALS:
            ws["A"+str(i)] = date
            ws["C"+str(i)] = self.format_money(self.TOTALS[date])
            i += 1
        self.WB.save(self.SPREADSHEET_NAME)
        
    def wrapup(self):
        self.WB.save(self.SPREADSHEET_NAME)
        self.WB.close()
    
        
